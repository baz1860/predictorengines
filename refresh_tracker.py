#!/usr/bin/env python3
"""Rewrite the 'Project Ledger' sheet in Betting Tracker.xlsx from
data/ledger.csv. Run by update.sh after settlement; never touches the
'My Bets' or 'Dashboard' sheets. Fails soft if openpyxl is missing."""
import sys
from pathlib import Path

HERE = Path(__file__).parent
XLSX = HERE / "Betting Tracker.xlsx"
LEDGER = HERE / "data" / "ledger.csv"

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError as e:
    sys.exit(f"   tracker refresh skipped (missing package: {e.name})")

if not XLSX.exists() or not LEDGER.exists():
    sys.exit("   tracker refresh skipped (workbook or ledger not found)")

GBP = '£#,##0.00;(£#,##0.00);"-"'
led = pd.read_csv(LEDGER)

# Closing line value per bet (M3). Blank when no odds_history.csv snapshots exist.
try:
    from clv import compute_clv
    led["clv"] = compute_clv(led)
except Exception:
    led["clv"] = float("nan")

wb = load_workbook(XLSX)
pl = wb["Project Ledger"]

# clear old data rows (keep header row and the note in M1)
if pl.max_row > 1:
    pl.delete_rows(2, pl.max_row - 1)
cols = list(led.columns)
# ensure the clv column has a header (col 12; the note lives in col 13/M1)
if "clv" in cols:
    pl.cell(row=1, column=cols.index("clv") + 1, value="clv").font = \
        Font(name="Arial", size=10, bold=True)
for i, row in led.iterrows():
    for c, col in enumerate(cols, 1):
        v = row[col]
        cell = pl.cell(row=i + 2, column=c, value=(None if pd.isna(v) else v))
        cell.font = Font(name="Arial", size=10)
        if col in ("stake", "pnl", "bankroll_after"):
            cell.number_format = GBP
        elif col == "odds":
            cell.number_format = "0.00"
        elif col == "clv":
            cell.number_format = "+0.0%;-0.0%"

wb.calculation.fullCalcOnLoad = True   # Excel recalcs Dashboard on open
wb.save(XLSX)
print(f"   Betting Tracker.xlsx refreshed ({len(led)} ledger rows)")
